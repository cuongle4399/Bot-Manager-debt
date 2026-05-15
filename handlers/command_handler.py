import os
import re
import sqlite3
import asyncio
import io
import psutil
import platform
import sys
import time
from datetime import datetime
from telegram import Update
from telegram.ext import ContextTypes
from services.debt_service import calculate_group_debts, get_my_debts
from utils.parser import format_currency
from config import OWNER_ID
from database.db_manager import (
    get_all_groups, 
    find_user_id_by_username, 
    get_user_id_or_pseudo,
    get_debts_in_group, 
    get_transactions_by_user,
    delete_transaction, 
    delete_transactions_by_message,
    save_transaction,
    clear_group_data
)
from openpyxl import Workbook
from openpyxl.styles import Font

BOT_START_TIME = time.time()

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    chat_type = update.message.chat.type
    
    if user_id == OWNER_ID:
        if chat_type != "private":
            group_id = update.message.chat.id
            message_id = update.message.message_id
            
            # 1. Thực hiện reset toàn bộ dữ liệu của group trong DB
            deleted_count = clear_group_data(group_id)
            
            # 2. Thực hiện dọn dẹp tin nhắn (Quét ngược 200 tin nhắn gần nhất)
            # Xóa chính tin nhắn lệnh !start
            try: await context.bot.delete_message(group_id, message_id)
            except: pass
            
            cleared_msgs = 0
            # Tối ưu: Chia nhỏ 500 tin nhắn thành các cụm 30 để xóa song song (Parallel deletion)
            chunk_size = 30
            for i in range(1, 501, chunk_size):
                # Tạo danh sách các tác vụ xóa trong cụm này
                tasks = []
                for j in range(i, min(i + chunk_size, 501)):
                    tasks.append(context.bot.delete_message(group_id, message_id - j))
                
                # Thực hiện xóa song song cả cụm
                results = await asyncio.gather(*tasks, return_exceptions=True)
                
                # Đếm số tin nhắn xóa thành công (không có lỗi)
                cleared_msgs += sum(1 for r in results if r is True)
                
                # Nghỉ ngắn giữa các cụm để tránh bị Telegram đánh dấu spam (Flood limit)
                await asyncio.sleep(0.4)
            
            msg = (
                "⚠️ **HỆ THỐNG ĐÃ ĐƯỢC RESET TOÀN DIỆN**\n\n"
                f"• **Dữ liệu nợ:** Đã xóa sạch `{deleted_count}` bản ghi.\n"
                f"• **Tin nhắn:** Đã dọn dẹp `{cleared_msgs}` tin nhắn gần đây.\n\n"
                "Sư phụ đã dọn dẹp sạch sẽ, chúc mọi người bắt đầu chu kỳ mới vui vẻ! 💸"
            )
            await update.message.reply_text(msg, parse_mode="Markdown")
        else:
            await update.message.reply_text("Chào Sư phụ! Hệ thống đang hoạt động ổn định. Gõ !admin để xem các lệnh quản trị.")
    else:
        await update.message.reply_text("Bot Quản Lý Công Nợ đã sẵn sàng. Gõ !help để xem hướng dẫn.")

async def myid_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(f"ID Telegram của bạn: `{update.message.from_user.id}`", parse_mode="Markdown")

async def idgroups_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.from_user.id != OWNER_ID:
        return
        
    group_ids = get_all_groups()
    if not group_ids:
        await update.message.reply_text("Chưa có danh sách nhóm nào trong bộ nhớ.")
        return
        
    res = "📊 **DANH SÁCH GROUP ID ĐANG HOẠT ĐỘNG**\n\n"
    for gid in group_ids:
        res += f"• `{gid}`\n"
    await update.message.reply_text(res, parse_mode="Markdown")

async def help_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    help_text = """
📜 **HƯỚNG DẪN BIẾT DÙNG BOT QUẢN LÝ NỢ**

📝 **CÁCH GHI NỢ:**
• `@user -50k [lý do]` : Bạn cho @user vay (Ghi nợ cho họ)
• `@user +50k [lý do]` : Bạn vay từ @user (Ghi nợ cho chính mình)
• `@u1 @u2 -50k [lý do]` : Ghi nợ cho nhiều người cùng lúc

📊 **TRA CỨU THÔNG TIN:**
• `!no` : Xem danh sách các khoản nợ của chính bạn
• `!no all` : Xem tổng hợp công nợ của tất cả thành viên
• `!ls` : Xem lịch sử giao dịch gần đây của bạn
• `!ls @user` : Xem lịch sử nợ chi tiết giữa bạn và @user
• `!myid` : Lấy ID Telegram của bạn

📢 **NHẮC NỢ & THANH TOÁN:**
• `!nhacno` : Gửi tin nhắn nhắc nhở tất cả những ai đang nợ bạn
• `!allpaid @user` : Xác nhận @user đã trả hết sạch nợ cho bạn
• `!undo` : Hoàn tác (xóa) đơn nợ vừa ghi 
   _(Hoặc Reply tin nhắn nợ bất kỳ và gõ !undo để xóa đơn đó)_

📥 **TIỆN ÍCH:**
• `!export` : Xuất file Excel lịch sử nợ cá nhân để đối soát

💡 *Lưu ý: Bạn chỉ có thể hoàn tác hoặc tất toán các khoản nợ liên quan đến chính mình.*
"""
    await update.message.reply_text(help_text, parse_mode="Markdown")

async def admin_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    if user_id != OWNER_ID:
        await update.message.reply_text("❌ Lệnh này chỉ dành cho Sư phụ Bot.")
        return
        
    admin_text = """
👑 **BẢNG ĐIỀU KHIỂN SƯ PHỤ (OWNER PANEL)** 👑
━━━━━━━━━━━━━━━━━━━━━━━━━━

🛠 **QUẢN LÝ DỮ LIỆU & HOẠT ĐỘNG**
• `!clear [số]` : 🧹 Dọn dẹp hàng loạt tin nhắn rác.
• `!allpaid @A @B` : 🤝 Tất toán nợ hộ cho 2 user bất kỳ.
• `!exportno` : 📊 Xuất file Excel tổng hợp công nợ cả nhóm.
• `!idgroups` : 🕵️ Quét ID các nhóm Bot đang hoạt động.
• `!no @user` : 🔍 Kiểm tra nợ của một thành viên bất kỳ.

⚙️ **QUẢN TRỊ HỆ THỐNG & MÁY CHỦ**
• `!start` : ⚠️ Xóa TOÀN BỘ nợ & dọn tin nhắn.
• `!ping` : 🏓 Check health hệ thống.
• `!rstbot` : 🔄 Force restart Bot.

💡 _Chỉ Sư phụ có ID trùng khớp trong cấu hình mới sử dụng được các lệnh này._
"""
    await update.message.reply_text(admin_text, parse_mode="Markdown")

async def no_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.chat.type == "private": return
    group_id, user_id = update.message.chat.id, update.message.from_user.id
    tags = re.findall(r"@(\w+)", update.message.text)
    
    is_owner = (user_id == OWNER_ID)

    if len(tags) >= 2:
        u1_name, u2_name = tags[0], tags[1]
        u1_id, u2_id = get_user_id_or_pseudo(u1_name), get_user_id_or_pseudo(u2_name)
        if not u1_id or not u2_id:
            await update.message.reply_text("Không tìm thấy dữ liệu.")
            return
        pair_debts, _ = calculate_group_debts(group_id)
        key = tuple(sorted((u1_id, u2_id)))
        amount = pair_debts.get(key, 0)
        if u1_id > u2_id: amount = -amount
        if amount == 0: await update.message.reply_text(f"@{u1_name} và @{u2_name} đang hòa nhau.")
        elif amount > 0: await update.message.reply_text(f"@{u2_name} đang nợ @{u1_name}: {format_currency(amount)}")
        else: await update.message.reply_text(f"@{u1_name} đang nợ @{u2_name}: {format_currency(abs(amount))}")
        return

    if len(tags) == 1 and is_owner:
        t_name = tags[0]
        t_id = get_user_id_or_pseudo(t_name)
        if not t_id: return
        owe_them, they_owe = get_my_debts(t_id, group_id)
        res = f"📊 **DANH SÁCH NỢ CỦA @{t_name.upper()}**\n\n"
        has = False
        for i in they_owe: res += f"• {t_name} nợ **{i['name']}**: `{format_currency(i['amount'])}`\n"; has = True
        for i in owe_them: res += f"• **{i['name']}** nợ {t_name}: `{format_currency(i['amount'])}`\n"; has = True
        if not has: res = f"✅ **@{t_name}** hiện không có nợ."
        await update.message.reply_text(res, parse_mode="Markdown"); return

    if context.args and context.args[0] == "all":
        pair_debts, user_names = calculate_group_debts(group_id)
        if not pair_debts:
            await update.message.reply_text("Nhom hien khong co no.")
            return
        res = "📋 **TỔNG HỢP CÔNG NỢ TOÀN NHÓM**\n\n"
        index = 1
        for (u1, u2), amount in pair_debts.items():
            if amount == 0: continue
            n1, n2 = user_names.get(u1, f"User {u1}"), user_names.get(u2, f"User {u2}")
            if amount > 0: res += f"{index}. **{n2}** nợ **{n1}**: `{format_currency(amount)}`\n"
            else: res += f"{index}. **{n1}** nợ **{n2}**: `{format_currency(abs(amount))}`\n"
            index += 1
        await update.message.reply_text(res, parse_mode="Markdown")
    else:
        owe_them, they_owe = get_my_debts(user_id, group_id)
        title = "📊 **DANH SÁCH NỢ CỦA BẠN**\n\n"
        res = ""
        for i in they_owe: res += f"• Bạn đang nợ **{i['name']}**: `{format_currency(i['amount'])}`\n"
        for i in owe_them: res += f"• **{i['name']}** đang nợ bạn: `{format_currency(i['amount'])}`\n"
        
        if not res:
            res = "✅ Hiện tại bạn không có nợ nào."
        else:
            res = title + res
        await update.message.reply_text(res, parse_mode="Markdown")

async def lichsu_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.chat.type == "private":
        await update.message.reply_text("❌ Lệnh này chỉ khả dụng trong Group chat.")
        return
    
    group_id, user_id = update.message.chat.id, update.message.from_user.id
    target_username = context.args[0].replace("@", "") if context.args else None
    
    # Lấy giao dịch liên quan đến mình bằng SQL để tối ưu hiệu suất
    my_raw_txs = get_transactions_by_user(group_id, user_id)

    if target_username:
        target_id = get_user_id_or_pseudo(target_username)
        # Lọc giao dịch giữa mình và Target
        filtered_raw = [t for t in my_raw_txs if t['creditor_id'] == target_id or t['debtor_id'] == target_id]
        
        last_settled_id = 0
        for t in filtered_raw:
            if "[TẤT TOÁN]" in t['reason']:
                last_settled_id = max(last_settled_id, t['id'])
        
        my_txs = [t for t in filtered_raw if t['id'] >= last_settled_id][:20]
        title = f"📜 **LỊCH SỬ NỢ VỚI @{target_username.upper()}**\n\n"
    else:
        peers_last_settled = {}
        for t in my_raw_txs:
            if "[TẤT TOÁN]" in t['reason']:
                peer = t['creditor_name'] if t['debtor_id'] == user_id else t['debtor_name']
                peers_last_settled[peer.lower()] = max(peers_last_settled.get(peer.lower(), 0), t['id'])
        
        filtered_txs = []
        for t in my_raw_txs:
            peer = t['creditor_name'] if t['debtor_id'] == user_id else t['debtor_name']
            if t['id'] >= peers_last_settled.get(peer.lower(), 0):
                filtered_txs.append(t)

        limit = 50 if (context.args and context.args[0] == "all") else 15
        my_txs = filtered_txs[:limit]
        title = "📜 **LỊCH SỬ GIAO DỊCH**\n\n"

    if not my_txs:
        await update.message.reply_text("Không tìm thấy dữ liệu lịch sử.")
        return

    res = title
    for t in my_txs:
        date_str = t['created_at'].split(".")[0] if isinstance(t['created_at'], str) else t['created_at'].strftime("%H:%M %d/%m")
        res += f"• `{date_str}` | **{t['debtor_name']}** nợ **{t['creditor_name']}** | {t['reason']} : `{format_currency(t['amount'])}`\n"
    
    await update.message.reply_text(res, parse_mode="Markdown")

async def undo_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.chat.type == "private":
        await update.message.reply_text("❌ Lệnh này chỉ khả dụng trong Group chat.")
        return

    group_id = update.message.chat.id
    user_id = update.message.from_user.id
    
    # 1. Trường hợp Undo theo Reply tin nhắn
    if update.message.reply_to_message:
        replied_msg_id = update.message.reply_to_message.message_id
        
        # Xóa tất cả transactions có cùng message_id đó (đề phòng tag nhiều người)
        deleted_count = delete_transactions_by_message(replied_msg_id, user_id)
        
        if deleted_count > 0:
            await update.message.reply_text(f"✅ **HOÀN TÁC THÀNH CÔNG**\nĐã xóa `{deleted_count}` đơn nợ liên quan đến tin nhắn này.", parse_mode="Markdown")
        else:
            await update.message.reply_text("❌ Không tìm thấy giao dịch hợp lệ do bạn tạo để hoàn tác.")
        return

    # 2. Trường hợp Undo lệnh cuối cùng (không reply)
    txs = get_debts_in_group(group_id)
    if not txs:
        await update.message.reply_text("Chưa có giao dịch nào trong nhóm này.")
        return

    # Tìm message_id của giao dịch gần nhất do user này tạo
    last_msg_id = None
    last_tx_info = None
    
    for t in sorted(txs, key=lambda x: x['id'], reverse=True):
        if t['created_by'] == user_id:
            last_msg_id = t['message_id']
            last_tx_info = t
            break
            
    if not last_msg_id:
        await update.message.reply_text("Bạn chưa tạo giao dịch nào gần đây để hoàn tác.")
        return
        
    # Xóa tất cả transactions có chung message_id đó
    deleted_count = delete_transactions_by_message(last_msg_id, user_id)
    
    if deleted_count > 0:
        msg = f"✅ **HOÀN TÁC THÀNH CÔNG**\n"
        msg += f"• Đã xóa `{deleted_count}` đơn nợ từ lệnh cuối cùng.\n"
        if deleted_count == 1:
            msg += f"• Nội dung: **{last_tx_info['debtor_name']}** nợ **{last_tx_info['creditor_name']}** (`{format_currency(last_tx_info['amount'])}`)\n"
        msg += f"• Lý do: {last_tx_info['reason']}"
        await update.message.reply_text(msg, parse_mode="Markdown")
    else:
        await update.message.reply_text("❌ Lỗi: Không thể hoàn tác giao dịch.")

async def nhacno_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.chat.type == "private":
        await update.message.reply_text("❌ Lệnh này chỉ khả dụng trong Group chat.")
        return
    group_id = update.message.chat.id
    user_id = update.message.from_user.id

    owe_me, _ = get_my_debts(user_id, group_id)
    
    if not owe_me:
        await update.message.reply_text("✅ Tuyệt vời! Hiện tại không có ai nợ bạn.")
        return
        
    res = "📢 **THÔNG BÁO NHẮC NỢ**\n\n"
    for item in owe_me:
        res += f"• @{item['name']} đang nợ: **{format_currency(item['amount'])}**\n"
    res += "\n💡 _Vui lòng nhắn tin riêng hoặc chuyển khoản để xóa nợ!_"
    await update.message.reply_text(res, parse_mode="Markdown")

async def allpaid_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.chat.type == "private":
        await update.message.reply_text("❌ Lệnh này chỉ khả dụng trong Group chat.")
        return
    tags = re.findall(r"@(\w+)", update.message.text)
    if not tags:
        await update.message.reply_text("Vui lòng tag người nợ.", parse_mode="Markdown")
        return
        
    group_id, user_id = update.message.chat.id, update.message.from_user.id
    is_owner = (user_id == OWNER_ID)

    if len(tags) >= 2:
        if not is_owner:
            await update.message.reply_text("❌ Lệnh này chỉ dành cho Sư phụ!")
            return
        u1_name, u2_name = tags[0], tags[1]
        u1_id, u2_id = find_user_id_by_username(u1_name), find_user_id_by_username(u2_name)
        if not u1_id or not u2_id:
            await update.message.reply_text("❌ Khong tim thay du lieu.")
            return
        pair_debts, user_names = calculate_group_debts(group_id)
        ids = sorted([u1_id, u2_id]); key = (ids[0], ids[1])
        net_amount = pair_debts.get(key, 0)
        if net_amount == 0:
            await update.message.reply_text(f"✅ @{u1_name} va @{u2_name} khong co no nhau.")
            return
        if net_amount > 0:
            debtor_id, debtor_name = ids[1], user_names.get(ids[1], u2_name)
            creditor_id, creditor_name = ids[0], user_names.get(ids[0], u1_name)
            amount = net_amount
        else:
            debtor_id, debtor_name = ids[0], user_names.get(ids[0], u1_name)
            creditor_id, creditor_name = ids[1], user_names.get(ids[1], u2_name)
            amount = abs(net_amount)

        save_transaction(group_id, {"id": debtor_id, "name": debtor_name}, {"id": creditor_id, "name": creditor_name},
                         amount, f"[ADMIN TẤT TOÁN] @{u1_name} & @{u2_name}", update.message.text, user_id, update.message.message_id)
        await update.message.reply_text(f"✅ [ADMIN] Đã tất toán giữa @{u1_name} và @{u2_name}.")
        return

    target_username = tags[0]
    target_id = get_user_id_or_pseudo(target_username)
    if target_id is None:
        await update.message.reply_text("❌ Không tìm thấy dữ liệu.")
        return

    pair_debts, _ = calculate_group_debts(group_id)
    ids = sorted([user_id, target_id]); key = (ids[0], ids[1])
    net_amount = pair_debts.get(key, 0)
    debt_to_me = net_amount if user_id == ids[0] else -net_amount
    if debt_to_me <= 0:
        await update.message.reply_text(f"@{target_username} khong no ban.")
        return

    save_transaction(group_id, {"id": target_id, "name": target_username}, {"id": user_id, "name": update.message.from_user.full_name},
                     debt_to_me, f"[TẤT TOÁN] @{target_username}", update.message.text, user_id, update.message.message_id)
    await update.message.reply_text(f"✅ Đã tất toán! @{target_username} trả {format_currency(debt_to_me)}.")

async def export_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    group_id = update.message.chat.id
    txs = get_debts_in_group(group_id)
    
    # Logic lọc y hệt như !ls (chỉ lấy từ lần tất toán gần nhất của mỗi cặp)
    peers_last_settled = {}
    my_raw_txs = [t for t in txs if t['creditor_id'] == user_id or t['debtor_id'] == user_id]
    for t in my_raw_txs:
        if "[TẤT TOÁN]" in t['reason']:
            peer = t['creditor_name'] if t['debtor_id'] == user_id else t['debtor_name']
            peers_last_settled[peer.lower()] = max(peers_last_settled.get(peer.lower(), 0), t['id'])
    
    my_txs = []
    for t in my_raw_txs:
        peer = t['creditor_name'] if t['debtor_id'] == user_id else t['debtor_name']
        if t['id'] >= peers_last_settled.get(peer.lower(), 0):
            my_txs.append(t)
            
    if not my_txs:
        await update.message.reply_text("Không có dữ liệu nợ hiện tại để xuất.")
        return
        
    wb = Workbook(); ws = wb.active; ws.title = "Lich Su No"
    ws.append(["ID", "Thời Gian", "Người Cho Nợ", "Người Nợ", "Số Tiền", "Lý Do"])
    for cell in ws[1]: cell.font = Font(bold=True)
    
    # Sắp xếp ID từ nhỏ đến lớn trong Excel để dễ theo dõi
    my_txs.sort(key=lambda x: x['id'])
    
    for t in my_txs:
        date_str = t['created_at'].split(".")[0] if isinstance(t['created_at'], str) else t['created_at'].strftime("%Y-%m-%d %H:%M")
        ws.append([t['id'], date_str, t['creditor_name'], t['debtor_name'], t['amount'], t['reason']])
        
    byte_io = io.BytesIO(); wb.save(byte_io); byte_io.seek(0); byte_io.name = "lich_su_no.xlsx"
    await update.message.reply_document(document=byte_io, caption="📊 File Excel lịch sử nợ hiện tại (sau tất toán).")

async def exportno_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    group_id = update.message.chat.id
    pair_debts, user_names = calculate_group_debts(group_id)
    actual_debts = {k: v for k, v in pair_debts.items() if v != 0}
    if not actual_debts:
        await update.message.reply_text("Nhom hien tai khong co no.")
        return
    wb = Workbook(); ws = wb.active; ws.title = "Tong Hop No"
    ws.append(["STT", "Người Nợ", "Người Cho Nợ", "Số Tiền Nợ"])
    for cell in ws[1]: cell.font = Font(bold=True)
    index = 1
    for (u1, u2), amount in actual_debts.items():
        n1, n2 = user_names.get(u1, f"User {u1}"), user_names.get(u2, f"User {u2}")
        if amount > 0: ws.append([index, n2, n1, amount])
        else: ws.append([index, n1, n2, abs(amount)])
        index += 1
    byte_io = io.BytesIO(); wb.save(byte_io); byte_io.seek(0); byte_io.name = "tong_hop_no.xlsx"
    await update.message.reply_document(document=byte_io, caption="📋 Bản tóm tắt công nợ.")

async def ping_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    # System Info
    cpu_usage = psutil.cpu_percent(interval=None)
    ram = psutil.virtual_memory()
    rom = psutil.disk_usage('/')
    
    # Uptime
    uptime_seconds = int(time.time() - BOT_START_TIME)
    days, rem = divmod(uptime_seconds, 86400)
    hours, rem = divmod(rem, 3600)
    minutes, seconds = divmod(rem, 60)
    uptime_str = f"{days}d {hours}h {minutes}m {seconds}s" if days > 0 else f"{hours}h {minutes}m {seconds}s"

    msg = "🏓 **PONG! SYSTEM STATUS**\n\n"
    msg += f"🖥 **Hệ điều hành:** `{platform.system()} {platform.release()}`\n"
    msg += f"⚙️ **CPU:** `{cpu_usage}%`\n"
    msg += f"💾 **RAM:** `{ram.percent}%` ({ram.used // (1024**2)}MB / {ram.total // (1024**2)}MB)\n"
    msg += f"🗄 **Disk:** `{rom.percent}%` ({rom.used // (1024**3)}GB / {rom.total // (1024**3)}GB)\n"
    msg += f"⏰ **Bot Uptime:** `{uptime_str}`\n"
    msg += f"📡 **Ping:** `{context.bot.token[:5]}...` (Stable)"
    
    await update.message.reply_text(msg, parse_mode="Markdown")

async def rstbot_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.message.from_user.id != OWNER_ID:
        await update.message.reply_text("❌ Lệnh này chỉ dành cho Sư phụ!")
        return
    
    await update.message.reply_text("🔄 **Đang khởi động lại Bot...**\nVui lòng đợi giây lát.")
    
    # Đảm bảo đóng kết nối DB (nếu cần) và chuẩn bị restart
    if getattr(sys, 'frozen', False):
        # Trường hợp chạy file .exe (đã đóng gói)
        import subprocess
        
        # Bỏ _MEIPASS2 để PyInstaller không tái sử dụng thư mục temp đang bị process cũ khóa
        env = os.environ.copy()
        if '_MEIPASS2' in env:
            del env['_MEIPASS2']
            
        CREATE_NO_WINDOW = 0x08000000
        subprocess.Popen([sys.executable] + sys.argv[1:], env=env, creationflags=CREATE_NO_WINDOW)
        os._exit(0)
    else:
        # Trường hợp chạy script .py
        os.execv(sys.executable, ['python'] + sys.argv)

async def clear_command(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    if user_id != OWNER_ID:
        await update.message.reply_text("❌ Lệnh này chỉ dành cho Sư phụ!")
        return
    amount = int(context.args[0]) if context.args else 10
    # Cho phép dọn dẹp tối đa 500 tin nhắn một lần
    if amount > 500: amount = 500
    
    chat_id, message_id = update.message.chat.id, update.message.message_id
    
    # Xóa chính tin nhắn lệnh !clear
    try: await context.bot.delete_message(chat_id, message_id)
    except: pass

    deleted = 0
    scan_range = int(amount * 1.2)
    chunk_size = 25 # Cụm nhỏ hơn cho lệnh clear để mượt hơn
    
    for i in range(1, scan_range + 1, chunk_size):
        if deleted >= amount: break
        
        tasks = []
        for j in range(i, min(i + chunk_size, scan_range + 1)):
            if deleted + len(tasks) >= amount: break
            tasks.append(context.bot.delete_message(chat_id, message_id - j))
        
        results = await asyncio.gather(*tasks, return_exceptions=True)
        success_in_chunk = sum(1 for r in results if r is True)
        deleted += success_in_chunk
        
        await asyncio.sleep(0.3)
    info_msg = await context.bot.send_message(chat_id, f"🧹 Đã dọn dẹp {deleted} tin nhắn!")
    await asyncio.sleep(3)
    await context.bot.delete_message(chat_id, info_msg.message_id)
